import argparse
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import time
import glob
import os, sys
import unidecode

from myFunctions import *


from argparse import ArgumentParser
import sys
from sys import argv

# Where the files that go into script are kept
path = "C:/Users/sakiikas/Documents/ScriptFiles_TEST/RC_SIS_FILES"
os.chdir("C:/Users/sakiikas/Documents/ScriptFiles_TEST/RC_SIS_FILES")


# On Lynda's Computer:
# path = "W:/Records/LyndaScript-master/RC_SIS_Files"
# os.chdir("W:/Records/LyndaScript-master/RC_SIS_Files")



RC_download_file = (glob.glob("*RC*")[0])
SIS_download_file = (glob.glob("*SIS*")[0])


workbook1 = (RC_download_file)
workbook2 = (SIS_download_file)
workbook3 = ('combined_workbook.xlsx')

wb1 = load_workbook(workbook1)
ws1 = wb1.active

wb2 = load_workbook(workbook2)
ws2 = wb2.active

wb3 = Workbook()
ws3_all_info = wb3.active

copy_paste_lookupid(ws1, ws3_all_info)
copy_paste_initial_info(ws1, ws3_all_info)
copy_paste_other_info(ws1, ws3_all_info)
# if "RC" in dirs[0]:
#     first_file = 'Ruffalo Cody'
#     second_file = 'Registrar: SIS Import'
# else:
#     first_file = 'Registrar: SIS Import'
#     second_file = 'Ruffalo Cody'
create_source_column(RC_download_file, ws3_all_info)

length = len(ws3_all_info['A']) + 1
append_second_worksheet_initial_info(ws2, ws3_all_info)
append_second_worksheet_other_info(ws2, ws3_all_info, length, SIS_download_file)

categorize_emails(ws3_all_info)
format_phone_number(ws3_all_info)
remove_accents(ws3_all_info)
format_country(ws3_all_info)
format_postal_code(ws3_all_info)
format_first_row(ws3_all_info)
format_non_initium_address(ws3_all_info)

# Where the new combined file will be saved to
wb3.save("C:/Users/sakiikas/Documents/ScriptFiles_TEST/RC_SIS_FILES/combined_workbook.xlsx")

# wb3.save("W:/Records/LyndaScript-master/RC_SIS_Files/combined_workbook.xlsx")

