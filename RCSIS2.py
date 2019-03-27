import argparse
from openpyxl import load_workbook, Workbook
import time
import os, sys
from myFunctions import *

# Changes the directory so you are where the created combined file was saved to
os.chdir("C:/Users/sakiikas/Documents/ScriptFiles_TEST/RC_SIS_FILES")

# os.chdir("W:/Records/LyndaScript-master/RC_SIS_Files")

clock = 'Combined_Workbook.xlsx'

workbook1 = ('Combined_Workbook.xlsx')
wb1 = load_workbook(workbook1)
ws3_all_info = wb1.active

wb4 = Workbook()
initium_Canada_info = wb4.active

wb5 = Workbook()
initium_USA_info = wb5.active


# Hypothetically everything has been fixed. Now create these new files
copy_paste_to_initium_file(ws3_all_info, initium_Canada_info, "Canada")
copy_paste_to_initium_file(ws3_all_info, initium_USA_info, "United States of America")


# TO DO: This works when I don't touch the newly created example3 file but doesn't hit hello kitty if i open file
for cell in ws3_all_info['R']:
    if cell.fill.fgColor.rgb == '00FFFF33' or cell.fill.fgColor.rgb == 'FFFFFF33':
        if cell.offset(row=0, column=1).value is not None:
            cell.fill = PatternFill(fill_type=None)


wb1.save("C:/Users/sakiikas/Documents/ScriptFiles_TEST/RC_SIS_FILES/Combined_Workbook.xlsx")
wb4.save("C:/Users/sakiikas/Documents/ScriptFiles_TEST/RC_SIS_FILES/RCSIS_Initium_Ready_CA.xlsx")
wb5.save("C:/Users/sakiikas/Documents/ScriptFiles_TEST/RC_SIS_FILES/RCSISInitium_Ready_USA.xlsx")

# wb1.save("W:/Records/LyndaScript-master/RC_SIS_Files/Combined_Workbook.xlsx")
# wb4.save("W:/Records/LyndaScript-master/RC_SIS_Files/RCSIS_Initium_Ready_CA.xlsx")
# wb5.save("W:/Records/LyndaScript-master/RC_SIS_Files/RCSISInitium_Ready_USA.xlsx")


