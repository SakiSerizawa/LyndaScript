import argparse
from openpyxl import load_workbook, Workbook
import time
import os, sys
from myFunctions import *

# Changes the directory so you are where the created combined file was saved to
os.chdir("../../../sakiikas/Documents/ScriptFiles_TEST/Folder1")

clock = 'combined_workbook.xlsx'

workbook1 = ('combined_workbook.xlsx')
wb1 = load_workbook(workbook1)
ws3_all_info = wb1.active

#
# timestr = time.strftime("%Y%m%d")
# workbook4 = ('Initium.xlsx', timestr)

wb4 = Workbook()
initium_Canada_info = wb4.active
initium_Canada_info.title = ("Canada")
initium_USA_info = wb4.create_sheet("USA")

# Hypothetically everything has been fixed. Now create these new files
copy_paste_to_initium_file(ws3_all_info, initium_Canada_info, "Canada")
copy_paste_to_initium_file(ws3_all_info, initium_USA_info, "United States of America")

# TO DO: This works when I don't touch the newly created example3 file but doesn't hit hello kitty if i open file
for cell in ws3_all_info['R']:
    if cell.fill.fgColor.rgb == '00FFFF33' or cell.fill.fgColor.rgb == 'FFFFFF33':
        if cell.offset(row=0, column=1).value is not None:
            cell.fill = PatternFill(fill_type=None)


wb1.save("C:/Users/sakiikas/Documents/ScriptFiles_TEST/Folder1/combined_workbook.xlsx")
wb4.save("C:/Users/sakiikas/Documents/ScriptFiles_TEST/Folder1/'Initium_Ready.xlsx")



