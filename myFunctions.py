from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import *
from openpyxl.worksheet.datavalidation import DataValidation
import unidecode
from stateAbbreviations import *
from emailHandles import *


def copy_paste_lookupid(sourcews, destws):
    """Copy in source sheets lookup id's into destination sheet."""
    for col in sourcews.iter_rows(min_col=1,max_col=1, min_row=2):
        for cell in col:
            destws[cell.coordinate].value = cell.value


def copy_paste_initial_info(source_ws, desti_ws):
    """Takes information of columns in source worksheet, from first name to state inclusive"""
    """Pastes into the destination worksheet"""
    c = 2
    for row in source_ws.iter_rows(min_row=2,min_col=column_index_from_string('D'),
                                   max_col=column_index_from_string('I')):
        for cell in row:
                desti_ws.cell(row=cell.row, column=c).value = cell.value
                c = c+1
        c = 2
    return 0


def copy_paste_other_info(sourcews, destws):
    """Loops through the source sheet, finds key columns and ccpies them to their respective desination worksheet"""
    for col in sourcews.columns:
        for cell in col:
            if cell.value == 'City' or cell.value == 'CITY':
                for cell in sourcews[get_column_letter(cell.column)]:
                    destws.cell(row=cell.row, column=column_index_from_string('I')).value\
                        = cell.value.title()
            elif cell.value == 'STATE' or cell.value == 'State':
                for cell in sourcews[get_column_letter(cell.column)]:
                    destws.cell(row=cell.row, column=column_index_from_string('J')).value = cell.value
            elif cell.value == 'ZIP' or cell.value == 'Postal_Code':
                for cell in sourcews[get_column_letter(cell.column)]:
                    destws.cell(row=cell.row, column=column_index_from_string('K')).value = cell.value
            elif cell.value == 'COUNTRY' or cell.value == 'Country':
                for cell in sourcews[get_column_letter(cell.column)]:
                    destws.cell(row=cell.row, column=column_index_from_string('L')).value = cell.value
            elif cell.value == 'EMAIL1' or cell.value == 'Email':
                for cell in sourcews[get_column_letter(cell.column)]:
                    destws.cell(row=cell.row, column=column_index_from_string('R')).value = cell.value
            elif cell.value == 'Phone' or cell.value == 'PHONE':
                for cell in sourcews[get_column_letter(cell.column)]:
                    destws.cell(row=cell.row, column=column_index_from_string('O')).value = cell.value
            elif cell.value == 'LastChangeDate' or cell.value == 'SIS_LAST_UPDATE_DATE':
                for cell in sourcews[get_column_letter(cell.column)]:
                    destws.cell(row=cell.row, column=column_index_from_string('U')).value = cell.value
    return 0


def create_source_column(first_file, destws):

    for cell in destws['V']:
        cell.value = "Ruffalo Cody"

    dv = DataValidation(type="list", formula1='"Registrar: SIS Import, Ruffalo Cody"', allow_blank=True)

    # Add the data-validation object to the worksheet
    destws.add_data_validation(dv)

    dv.add('V2:V1048576')

    return 0


# def append_lookup_id(source_ws, dest_ws):
#     """Appends the lookup id from the source sheet onto the destination sheet"""
#     # Creates list of items that need to be appended
#     lookupid_list = []
#     for col in source_ws.iter_cols(min_col=1,max_col=1, min_row=2):
#         for cell in col:
#             lookupid_list.append(cell.value)
#
#     # Appends the lookup ID from the second workbook onto the first
#     for col in dest_ws.iter_rows(max_col=1):
#         if col[-1].value == 'LOOKUP ID':
#             for data in lookupid_list:
#                 dest_ws.append([data])
#     return 0


def append_second_worksheet_initial_info(source_ws, target_ws):
    """Appends columns (from first name to state inclusive) from secondary source worksheet to target worksheet """
    """The try catch is to remove accents, but if the cell is blank, doesn't attempt to remove acccents (avoid error)"""
    length1 = len(target_ws['A']) + 1
    for col in source_ws.iter_rows(min_row=2,min_col=column_index_from_string('D'),max_col=column_index_from_string('I')):
            row = [None]*1 + [cell.value for cell in col]
            target_ws.append(row)


    # Copy in source sheets lookup id's into destination sheet
    for col in source_ws.iter_rows(min_col=1,max_col=1, min_row=2):
        for cell in col:
            target_ws.cell(row=length1,column=1).value = cell.value
            length1=length1+1

    return 0


def append_second_worksheet_other_info(source_ws,target_ws, length_OG, sis_file):
    """Loops through the source sheet, finds key columns and appends them to a destination worksheet"""
    # Look through each source sheet column
    for col in source_ws.columns:
        # Within in column, checks to see if cell is appropriate header we are looking for
        for cell in col:
            if cell.value == 'CITY' or cell.value == 'City':
                for col in source_ws.iter_cols(min_row=2, max_col=cell.column, min_col=cell.column):
                    for cell in col:
                        target_ws.cell(row=length, column=column_index_from_string('I')).value = cell.value.title()
                        length = length+1
            elif cell.value == 'STATE' or cell.value == 'State':
                for col in source_ws.iter_cols(min_row=2, max_col=cell.column, min_col=cell.column):
                    for cell in col:
                        try:
                            target_ws.cell(row=length, column=column_index_from_string('J')).value = cell.value
                            length = length+1
                        except:
                            target_ws.cell(row=length, column=column_index_from_string('J')).value = cell.value
                            length = length+1
            elif cell.value == 'ZIP' or cell.value == 'Postal_Code':
                for col in source_ws.iter_cols(min_row=2, max_col=cell.column, min_col=cell.column):
                    for cell in col:
                        target_ws.cell(row=length, column=column_index_from_string('K')).value = cell.value
                        length = length+1
            elif cell.value == 'COUNTRY' or cell.value == 'Country':
                for col in source_ws.iter_cols(min_row=2, max_col=cell.column, min_col=cell.column):
                    for cell in col:
                        target_ws.cell(row=length, column=column_index_from_string('L')).value = cell.value
                        length = length+1
            elif cell.value == 'EMAIL' or cell.value == 'Email' or cell.value == 'EMAIL1':
                for col in source_ws.iter_cols(min_row=2, max_col=cell.column, min_col=cell.column):
                    for cell in col:
                        target_ws.cell(row=length, column=column_index_from_string('R')).value = cell.value
                        length = length+1
            elif cell.value == 'PHONE' or cell.value == 'Phone':
                for col in source_ws.iter_cols(min_row=2, max_col=cell.column, min_col=cell.column):
                    for cell in col:
                        cell.number_format = "0000000"
                        target_ws.cell(row=length, column=column_index_from_string('O')).value = cell.value
                        length = length+1
            elif cell.value == 'LastChangeDate' or cell.value == 'SIS_LAST_UPDATE_DATE':
                for col in source_ws.iter_cols(min_row=2, max_col=cell.column, min_col=cell.column):
                    for cell in col:
                        target_ws.cell(row=length, column=column_index_from_string('U')).value = cell.value
                        target_ws.cell(row=length, column=column_index_from_string('V')).value = "Registrar: SIS Import"
                        length = length+1
        length = length_OG
    return 0





def categorize_emails(worksheet, chosen_column, datavalidation_location1, datavalidation_location2):
    """Looking through the email column R, determines email category and marks column next to it accordingly"""
    """A is Alumni, H Home, and B Business. Anything not categorized is highlighted yellow in the target worksheet"""
    for cell in worksheet[chosen_column]:
        cell.offset(row=0, column=2).value = 0
        if cell.value is None:
            cell.offset(row=0, column=2).value = None
            continue
        elif cell.value.endswith('alumni.ubc.ca'):
            (cell.offset(row=0, column=1).value) = 'A'
        elif cell.value.lower().endswith(hometuple):
            (cell.offset(row=0, column=1).value) = 'H'
        elif cell.value.lower().endswith(businesstuple):
            (cell.offset(row=0, column=1).value) = 'B'
        elif "@alumni" in cell.value:
            (cell.offset(row=0, column=1).value) = 'H'

        else:
            try:
                last_name_email = worksheet.cell(row=cell.row, column=column_index_from_string('D')).value.lower() + ".ca"
                last_name_email_USA = worksheet.cell(row=cell.row,
                                                 column=column_index_from_string('D')).value.lower() + ".com"
                last_name_email_USA2 = worksheet.cell(row=cell.row,
                                                 column=column_index_from_string('D')).value.lower() + ".us"
                # print(last_name_email)
                if cell.value.endswith(last_name_email) or cell.value.endswith(last_name_email_USA) or cell.value.endswith(last_name_email_USA2) :
                    (cell.offset(row=0, column=1).value) = 'H'
                else:
                    cell.fill = PatternFill(fgColor='FFFF33', fill_type = 'solid')
                    (cell.offset(row=0, column=1).value) = 'B'

            except:
                cell.fill = PatternFill(fgColor='FFFF33', fill_type = 'solid')
                (cell.offset(row=0, column=1).value) = 'B'

    # Creates a data validation (drop down) object
    dv = DataValidation(type="list", formula1='"H,B,O,A"', allow_blank=True)
    dv2 = DataValidation(type="list", formula1='"0,1"', allow_blank=True)

    # Add the data-validation object to the worksheet
    worksheet.add_data_validation(dv)
    worksheet.add_data_validation(dv2)
    dv.add(datavalidation_location1)
    dv2.add(datavalidation_location2)

    return 0

def format_phone_number(worksheet):
    """Formats phone number by removing extra spaces and unnecessary characters"""
    for col in worksheet.iter_rows(min_row=2, min_col=column_index_from_string('O'), max_col=column_index_from_string('O')):
        for cell in col:
            """Excel rounds integers longers than 15 digits hence the large value in if statement below"""
            if (type(cell.value) == float) or (type(cell.value) == int and cell.value > 100000000000000):
                    cell.fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
                    break
            elif cell.value is not None:
                phone = str(cell.value)
                cell.value = phone.replace('(', '').replace('-', '').replace(')', '').replace(' ', '').replace(' ', '').\
                    replace('#', '').replace('.', '').replace('+','').replace('=','')
                if len(cell.value) > 10:
                    if cell.value.startswith('1') and (
                            cell.offset(row=0, column=-3).value == "Canada" or cell.offset(row=0,
                                                                                            column=-3).value == "United States Of America"):
                        cell.value = cell.value.replace('1', '')
                    else:
                        for key in country_codes:
                            if cell.value.startswith(key):
                                try:
                                    if country_codes[key] == cell.offset(row=0, column=-4).value:
                                        cell.value = cell.value.replace(key, '')
                                    break
                                except:
                                    pass
                            else:
                                cell.fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
                                break
                if len(str(cell.value)) > 10:
                    cell.fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
                try:
                    cell.value = int(cell.value)
                except:
                    pass


    for cell in worksheet['P']:
        if (cell.offset(row=0, column=-1).value) is None:
            continue
        else:
            cell.value = 'H'
            (cell.offset(row=0, column=1).value) = 1
    # Creates a data validation (drop down) object
    dv = DataValidation(type="list", formula1='"H,B,C"', allow_blank=True)
    dv2 = DataValidation(type="list", formula1='"0,1"', allow_blank=True)

    # Add the data-validation object to the worksheet
    worksheet.add_data_validation(dv)
    worksheet.add_data_validation(dv2)

    dv.add('P2:P1048576')
    dv2.add('Q2:Q1048576')

    return 0

def remove_accents(worksheet):
    """Trys to remove accents from First/Middle/Last name, Address1234, City, and Country"""
    for row in worksheet.iter_rows(min_col=2, max_col=column_index_from_string('L')):
        for cell in row:
            try:
                if cell.value is not None and (type(cell.value) != int):
                    cell.value = unidecode.unidecode(cell.value.strip().title())
            except:
                # print(cell.value)
                continue
    """Trys to removes accents from the Country column"""
    for cell in worksheet['Q']:
        try:
            cell.value = unidecode.unidecode(cell.value)
        except:
            continue


def format_country(worksheet):
    """ Changes country format to a type that LINKS """
    for cell in worksheet['L']:
        if cell.value == 'CANA':
            cell.value = 'Canada'
        elif cell.value == 'USA' or cell.value == 'United States':
            cell.value = 'United States of America'
        elif cell.value == 'CAMP':
            cell.value = 'Canada'
            (cell.offset(row=0, column=-2).value) = 'BC'
            (cell.offset(row=0, column=-3).value) = 'Vancouver'
        elif cell.value in countryDictionary.keys():
            cell.value = countryDictionary.get(cell.value)
        if cell.value in popular_countries and cell.offset(row=0, column=-2).value is not None:
            try:
                (cell.offset(row=0, column=-2).value) = popular_countries[cell.value][cell.offset(row=0, column=-2).value]
            except:
                cell.fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
                cell.offset(row=0, column=-1).fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
                continue
        """Removes state (unecessary) if the country is European, Singapore, or Taiwan"""
        if cell.value in european_countries_and_singapore and cell.offset(row=0,column=-3).value is not None:
            cell.offset(row=0, column=-2).value = ''

        elif cell.offset(row=0,column=-2).value is None:
            cell.offset(row=0, column=-2).fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
            cell.offset(row=0, column=-3).fill = PatternFill(fgColor='FDAB9F', fill_type='solid')






    """Sets the address type and Address is Primary option"""
    for cell in worksheet['M']:
        cell.value = 'H'
        (cell.offset(row=0, column=1).value) = 0
    # Creates a data validation (drop down) object
    dv = DataValidation(type="list", formula1='"H,B,C"', allow_blank=True)
    dv2 = DataValidation(type="list", formula1='"0,1"', allow_blank=True)

    # Add the data-validation object to the worksheet
    worksheet.add_data_validation(dv)
    worksheet.add_data_validation(dv2)



    dv.add('M2:M1048576')
    dv2.add('N2:N1048576')

    return 0

def format_postal_code(worksheet, country_column,offset_value):
    """ Formats Canadian Postal Codes to be 3 characters, a space, and three characters.
    Formats American Postal Codes so that it's 5 characters , dash then four: 12345-5555 EXAMPLE
    Formats Japanese Postal Codes so that it's three characters, then a dash: 123-1234 EXAMPLE
    If the postal code is an incorrect format, flag as pink"""
    for cell in worksheet[country_column]:
        if cell.value == 'Canada' or cell.value == 'Province':
            postal_code = cell.offset(row=0, column=offset_value).value
            try:
                if cell.offset(row=0, column=offset_value).value is None or cell.value == 'Province':
                    pass
                elif (postal_code[3] != ' ' or not postal_code[3].isdigit()) and postal_code is not None:
                    cell.offset(row=0, column=offset_value ).value = postal_code[:3] + ' ' + postal_code[3:]
                    cell.offset(row=0, column=offset_value).value = cell.offset(row=0, column=offset_value).value.replace('  ', ' ')
            except Exception as e:
                cell.fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
                cell.offset(row=0, column=offset_value ).fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
        if cell.value == 'United States of America':
            zipcode = cell.offset(row=0, column=offset_value).value
            if type(zipcode) != int and '-' not in zipcode:
                if cell.offset(row=0, column=offset_value).value is not None:
                    cell.fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
                    cell.offset(row=0, column=offset_value ).fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
            else:
                temp_str_zip = str(zipcode)
                cell.offset(row=0, column=offset_value).value = temp_str_zip[:5] + '-' + temp_str_zip[5:]
                if '--' in cell.offset(row=0, column=-1.).value:
                    cell.offset(row=0, column=offset_value ).value = cell.offset(row=0, column=offset_value ).value.replace('-', '', 1)
        if cell.value == 'Japan':
            zipcode = cell.offset(row=0, column=offset_value).value
            temp_str_zip = str(zipcode)
            cell.offset(row=0, column=offset_value).value = temp_str_zip[:3] + '-' + temp_str_zip[3:]


    return 0



title_row = ["LOOKUPID", "FIRST_NAME", "MIDDLE_NAME", "LAST_NAME", "Street1", "Street2", "Street3", "Street4", "CITY",
             "STATE", "Postal_Code", "COUNTRY","Address Type", "Address is Primary", "Phone", "Phone Type",
             "Phone is Primary", "Email", "Email Type", "Email is Primary", "Last_UPDT", "Source"]


def format_first_row(worksheet):
    """Formats the first row of the freshsly formatted excel sheet to have the proper titles, referenced from
    title list title_row"""
    i = 0
    for row in worksheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.value = title_row[i]
            cell.fill = PatternFill(fgColor='FFFFFF')
            cell.font = Font(bold=True)
            i = i+1
    return 0


def format_address(worksheet):
    """Formats address to be in title format (ie: Japan vs JAPAN) while removing unecessary symbols"""
    """Removes the word Apartment from Canadian address"""
    for cell in worksheet['E']:
        try:
            address1 = str(cell.value).title()
            address2 = str(cell.offset(row=0, column=1).value).title()
            address3 = str(cell.offset(row=0, column=2).value).title()
            address4 = str(cell.offset(row=0, column=3).value).title()

            cell.value = address1.replace('(', '').replace(')', '').replace('.', ' ').replace('#', ' ').replace(',',' ')\
                .replace('None', '').replace(' - ', '-').replace('- ', '-').replace(' -', '-').replace('  ', ' ').replace('Th', 'th')
            cell.offset(row=0, column=1).value = address2.replace('(', '').replace(')', '').replace('.', ' ').replace(
                '#', ' ').replace(',', ' ').replace('None', '').replace('- ', '-').replace(' -', '-').replace('  ', ' ').replace('Th', 'th')
            cell.offset(row=0, column=2).value = address3.replace('(', '').replace(')', '').replace('.', ' ').replace(
                '#', ' ').replace(',', ' ').replace('None', '').replace('- ', '-').replace(' -', '-').replace('  ', ' ').replace('Th', 'th')
            cell.offset(row=0, column=3).value = address4.replace('(', '').replace(')', '').replace('.', ' ').replace(
                '#', ' ').replace(',', ' ').replace('None', '').replace('- ', '-').replace(' -', '-').replace('  ', ' ').replace('Th', 'th')
        except Exception as e:
            continue

        for key in unwanted_words:
            if cell.value.startswith(key) and cell.offset(row=0, column = 7).value == "Canada":
                cell.value = cell.value.replace(key, unwanted_words[key])
                break

    return 0



initium_title_row = ["LOOKUP ID", "Street1", "Street2", "Street3", "Street4", "CITY",
             "STATE", "Postal_Code", "COUNTRY"]


def copy_paste_to_initium_file(source_ws, desti_ws, country):
    """Places information into a new excel file based on country"""
    for cell in source_ws['L']:
        if cell.value == country:
            alumni_info = []
            alumni_info.append(source_ws.cell(row=cell.row, column=1).value)
            for col in source_ws.iter_cols(min_row=cell.row, max_row=cell.row, min_col=column_index_from_string('E'),
                                           max_col=column_index_from_string('L')):
                for cell in col:
                    alumni_info.append(cell.value)
            desti_ws.append(alumni_info)
    desti_ws.insert_rows(1)
    i = 0
    for row in desti_ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.value = initium_title_row[i]
            cell.fill = PatternFill(fgColor='FFFFFF')
            cell.font = Font(bold=True, color='FF0000')
            i = i+1
    for cell in desti_ws['A']:
        cell.fill = PatternFill(fgColor='FFFF66', fill_type='solid')
    return 0


def colour_worksheet(target_ws):
    """Fills in cells of the worksheet to red"""
    for rows in target_ws.rows:
        for cell in rows:
            cell.fill = PatternFill(fgColor='FCD5B4', fill_type='solid')
    for cell in target_ws[1]:
        cell.fill=PatternFill(fgColor='FFFFFF', fill_type='solid')
        cell.font = Font(bold=True)
    return 0


def replace_info(ws_source, info_list):
    """Similar to VLOOKUP, matches LOOKUPID from the initium results document and matches to LOOKUPID in source ws.
    If there is a match, replaces the information with information in info_list"""

    for cell in ws_source['A']:
        cell.value = str(cell.value)
        # print("CELL TYPE: ", type(cell.value))
        # print("INFO_LIST TYPE: ", type(info_list[0]))
        if cell.value == info_list[0]:
            # print(cell.value, info_list[0])
            c = 1
            for found_row_cell in ws_source[cell.row]:
                found_row_cell.fill = PatternFill(fgColor='FFFFFF')
            for found_row_cell in ws_source[cell.row]:
                found_row_cell.offset(row=0, column=4).value = info_list[c]
                c = c+1
                if c == len(info_list):
                    return
            return 0


def make_column_list(ws_source, list_of_keys):
    """Takes in list of column letters, stores values from those columns into a list, and creates tuples
    of the rows of information"""
    all_info = []
    for key in list_of_keys:
        info_from_column = []
        for cell in ws_source[key]:
            info_from_column.append(cell.value)
        all_info.append(info_from_column)
    return zip(*all_info)


def create_data_validation(dv_object, dest_ws, column_choice):

    # Add the data-validation object to the worksheet
    dest_ws.add_data_validation(dv_object)

    column_range = str(column_choice + "2:" + column_choice + "1048576")
    dv_object.add(column_range)

    return 0

def convert_from_string_to_int(worksheet, column):
    """Converts most LOOKUPID's back to integers to prevent warnings in excel (ex "this number is stored as string"""
    for cell in worksheet[column]:
        try:
            cell.value = int(cell.value)
        except:
            continue

    return 0

def delete_empty_rows(worksheet, column):
    """Delete rows that have a lookupID without an address"""
    for cell in worksheet[column]:
        if cell.value is None:
            worksheet.delete_rows(cell.row, 1)
    return 0

