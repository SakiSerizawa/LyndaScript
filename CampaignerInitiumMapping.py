from openpyxl import load_workbook, Workbook
import time
import glob
from myFunctions import *
import os, sys

os.chdir("C:/Users/sakiikas/Documents/ScriptFiles_TEST/Campaigner")
# os.chdir("W:/Records/LyndaScript-master/Campaigner Files")

"""Loads the two excel files, Contact Update Template and Initium Canada-Results"""
contact_update_template_file = (glob.glob("*Contact_Update_Template*")[0])
initium_results_file = (glob.glob("*Canada-Results*")[0])

workbook1 = (contact_update_template_file)
workbook2 = (initium_results_file)

wb1 = load_workbook(workbook1)
ws1 = wb1.active

wb2 = load_workbook(workbook2)
ws2 = wb2["Exact"]


"""Fills in cells of the worksheet to red"""
for rows in ws1.iter_rows(min_row=2, min_col=1, max_col=column_index_from_string('L')):
    for cell in rows:
        cell.fill = PatternFill(fgColor='FCD5B4', fill_type='solid')
for cell in ws1[1]:
    cell.fill = PatternFill(fgColor='FFFFFF', fill_type='solid')
    cell.font = Font(bold=True)

"""Changes cells to white when there is a match between Contact Update Template and Initium Results"""
for col in ws2.iter_cols(min_col=1, max_col=1,min_row=3):
        for cell in col:
            initium_info = [str(ws2.cell(row=cell.row, column=1).value), ws2.cell(row=cell.row, column=10).value,
                            ws2.cell(row=cell.row, column=11).value, ws2.cell(row=cell.row, column=12).value,
                            ws2.cell(row=cell.row, column=13).value, ws2.cell(row=cell.row, column=15).value,
                            ws2.cell(row=cell.row, column=16).value, ws2.cell(row=cell.row, column=17).value]
            replace_info(ws1, initium_info)

"""Converts most LOOKUPID's back to integers to prevent warnings in excel (ex "this number is stored as string"""
for cell in ws1['A']:
    try:
        cell.value = int(cell.value)
    except:
        continue


try:
    initium_results_file_USA = (glob.glob("*USA-Results*")[0])
    workbook3 = (initium_results_file_USA)
    wb3 = load_workbook(workbook3)
    ws3 = wb3["Exact"]

    for col in ws3.iter_cols(min_col=1, max_col=1,min_row=3):
            for cell in col:
                initium_info = [str(ws3.cell(row=cell.row, column=1).value), ws3.cell(row=cell.row, column=10).value,
                                ws3.cell(row=cell.row, column=11).value, ws3.cell(row=cell.row, column=12).value,
                                ws3.cell(row=cell.row, column=13).value, ws3.cell(row=cell.row, column=15).value,
                                ws3.cell(row=cell.row, column=16).value, ws3.cell(row=cell.row, column=17).value]
                replace_info(ws1, initium_info)

except:
    print("No Initium USA-Results File or Error Processing Initium USA-Results File")

"""Converts most LOOKUPID's back to integers to prevent warnings in excel (ex "this number is stored as string"""
for cell in ws1['A']:
    try:
        cell.value = int(cell.value)
    except:
        continue


"""Sets the Address Type and Address is Primary option to green"""
for cell in ws1['M']:
    cell.fill = PatternFill(fgColor="D8E4BC", fill_type="solid")
    cell.border = Border(left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'),
                         top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000'), )
    cell.offset(row=0, column=1).fill = PatternFill(fgColor="D8E4BC", fill_type="solid")
    cell.offset(row=0, column=1).border = Border(left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'),
                         top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000'), )


"""Sets the Phone Type and Phone is Primary option to orange"""
for cell in ws1['Q']:
    cell.fill = PatternFill(fgColor="FDE9D9", fill_type="solid")
    cell.border = Border(left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'),
                         top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000'), )
    cell.offset(row=0, column=1).fill = PatternFill(fgColor="FDE9D9", fill_type="solid")
    cell.offset(row=0, column=1).border = Border(left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'),
                         top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000'), )

"""Sets the Email Type and Email is Primary option to blue"""
for cell in ws1['T']:
    cell.fill = PatternFill(fgColor="B7DEE8", fill_type="solid")
    cell.border = Border(left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'),
                         top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000'), )
    cell.offset(row=0, column=1).fill = PatternFill(fgColor="B7DEE8", fill_type="solid")
    cell.offset(row=0, column=1).border = Border(left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'),
                         top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000'), )


"""Sets the Source column of the worksheet to Online Contact Update"""
for row in ws1.iter_rows(min_row=2, min_col=column_index_from_string('W'), max_col=column_index_from_string('W')):
    for cell in row:
        cell.value = 'Online Contact Update'
        cell.fill = PatternFill(fgColor="DCE6F1", fill_type="solid")
        cell.border = Border(left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'),
                             top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000'), )

wb1.save("Campaigner - Contact_Update_Template.xlsx")