from openpyxl import load_workbook, Workbook
import time
import glob
from myFunctions import *
import os, sys


os.chdir("../../../sakiikas/Documents/ScriptFiles_TEST/RC_SIS_FILES")
# os.chdir("W:/Records/LyndaScript-master/RC_SIS_Files")

initium_results_file = (glob.glob("*Canada-Results*")[0])


workbook1 = ('combined_workbook.xlsx')
workbook2 = (initium_results_file)

wb1 = load_workbook(workbook1)
ws1 = wb1.active

wb2 = load_workbook(workbook2)
ws2 = wb2["Exact"]


colour_worksheet(ws1)

for col in ws2.iter_cols(min_col=1, max_col=1,min_row=3):
        for cell in col:
            initium_info = [str(ws2.cell(row=cell.row, column=1).value), ws2.cell(row=cell.row, column=10).value,
                            ws2.cell(row=cell.row, column=11).value, ws2.cell(row=cell.row, column=12).value,
                            ws2.cell(row=cell.row, column=13).value, ws2.cell(row=cell.row, column=15).value,
                            ws2.cell(row=cell.row, column=16).value, ws2.cell(row=cell.row, column=17).value]
            replace_info(ws1, initium_info)

"""Converts most LOOKUPID's back to integers to prevent warnings in excel (ex "this number is stored as string)"""
for cell in ws1['A']:
    cell.value = str(cell.value)


try:
    initium_results_file_USA = (glob.glob("*US-Results*")[0])
    workbook3 = (initium_results_file_USA)
    wb3 = load_workbook(workbook3)
    ws3 = wb3["Exact"]

    for col in ws3.iter_cols(min_col=1, max_col=1,min_row=3):
            for cell in col:
                initium_info = [str(ws3.cell(row=cell.row, column=1).value), ws3.cell(row=cell.row, column=10).value,
                                ws3.cell(row=cell.row, column=11).value, ws3.cell(row=cell.row, column=12).value,
                                ws3.cell(row=cell.row, column=13).value, ws3.cell(row=cell.row, column=15).value,
                                ws3.cell(row=cell.row, column=16).value, ws3.cell(row=cell.row, column=17).value]
                replace_info(ws1, initium_info)

except Exception as e:
    print("No Initium USA-Results File or Error Processing Initium USA-Results File")

"""Converts most LOOKUPID's back to integers to prevent warnings in excel (ex "this number is stored as string"""
for cell in ws1['A']:
    try:
        cell.value = int(cell.value)
    except:
        continue



wb1.save('combined_workbook.xlsx')

