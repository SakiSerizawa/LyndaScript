from openpyxl import load_workbook, Workbook
from myFunctions import *
import os, sys
import csv
from openpyxl.worksheet.datavalidation import DataValidation

# This script is to be run after the cmt file has run through the LINKS Constituent Matching Tool
# It is assumed the user has copied back the results back into the Campaigner_workbook.xlsx file

#
# The purpose of this script is 1) Create a CommMailPreferences file based off the updated Campaigner_workbook.xlsx
# 2) to delete any rows of data that could not be matched (through matching tool or manually
# 3) Prepare the file for LINKS

os.chdir("C:/Users/sakiikas/Documents/ScriptFiles_TEST/Campaigner")
# os.chdir("W:/Records/LyndaScript-master/Campaigner Files")

workbook1 = ('Campaigner_Workbook.xlsx')
wb1 = load_workbook(workbook1, data_only=True)
ws1 = wb1.active

"""Creates the workbook CommMailPreferences"""
wb2 = Workbook()
ws2 = wb2.active

"""Creates LINKS ready file of those who would only like to receive TREK online"""


"""Extracts LOOKUPID, FIRST/MIDDLE/LAST NAME, EMAIL from Campaigner_workbook and puts them in new CommPrefUpdate file"""
extra_row_info = ["General Correspondence", " ", "AA - TREK Magazine"," ", "No", "M", "Alumni Affairs",
                  "Requested by constituent"]
                  #"Last_UPDT", "Alumni Association"]

for cellz in ws1['AT']:
    row_info = []
    if "online" in cellz.value:
        if "unable to locate" not in str(ws1.cell(row=cellz.row, column=2).value):
            row_info.append(ws1.cell(row=cellz.row, column=2).value)
            row_info.append(ws1.cell(row=cellz.row, column=column_index_from_string('R')).value)
            row_info.append(ws1.cell(row=cellz.row, column=4).value)
            row_info.append(ws1.cell(row=cellz.row, column=5).value)
            row_info.append(ws1.cell(row=cellz.row, column=6).value)
            for item in extra_row_info:
                row_info.append(item)
            row_info.append(int(ws1.cell(row=cellz.row, column=column_index_from_string('AW')).value.strftime("%Y%m%d")))
            row_info.append("Alumni Association")
            ws2.append(row_info)


"""Converts most LOOKUPID's back to integers to prevent warnings in excel (ex "this number is stored as string"""
for cell in ws2['A']:
    try:
        cell.value = int(cell.value)
    except:
        continue


"""Creates a properly formatted Title Row in the CommMailPreference workbook"""
comm_mail_preferences_title_row = ["LOOKUP ID", "EMAIL", "FIRST_NAME", "MIDDLE_NAME", "LAST_NAME", "MAIL Type", "Site",
                                   "Correspondence code", "Category", "Send Yes/No", "Send by", "Business Unit",
                                   "Comments", "Last_UPDT", "Source"]

ws2.insert_rows(1,1)
i=0
for row in ws2.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.value = comm_mail_preferences_title_row[i]
        cell.font = Font(bold=True, color='FF0000')
        i = i + 1

with open('CommMailPreferences.csv', 'w', newline="") as csvfile:
    c = csv.writer(csvfile)
    for r in ws2.rows:
        c.writerow([cell.value for cell in r])

#wb2.save("CommMailPreferences.csv")

"""New workbook: Contact Update Template"""
wb3= Workbook()
ws3 = wb3.active

"""New workbook: Canadian Campaigner Initium Ready file"""
wb4= Workbook()
ws4 = wb4.active

"""New workbook: USA Campaigner Initium Ready file"""
wb6= Workbook()
ws6 = wb6.active

column_list = ['B', 'D', 'E', 'F', 'J', 'K', 'L', 'M','N', 'O', 'I', 'Q', 'P', 'R', 'AW']

x = make_column_list(ws1, column_list)



information_from_excel = list(x)
maximum_rows = len(information_from_excel)
maximum_col = len(information_from_excel[0])


i = 0
for rows in ws3.iter_rows(max_row=maximum_rows, max_col=maximum_col):
    j = 0
    if information_from_excel[i][j] == "unable to locate":
        pass
    else:
        for cell in rows:
            cell.value = information_from_excel[i][j]
            j = j + 1
    i = i + 1
ws3.insert_cols(column_index_from_string('H'), 1)
ws3.insert_cols(column_index_from_string('M'), 2)

"""Sets the Address Type and Address is Primary option to newly created column and manipulates the colour"""
for cell in ws3['M']:
    cell.value = 'H'
    cell.offset(row=0, column=1).value = 0
ws3.insert_cols(column_index_from_string('Q'), 2)
ws3.insert_cols(column_index_from_string('T'), 2)


"""Sets the Phone Type and Phone is Primary option to newly created column and manipulates the colour"""
for cell in ws3['Q']:
    if cell.offset(row=0, column=-1).value is None:
        pass
    elif cell.offset(row=0,column=-2).value == "Home Cell Phone":
        cell.value = 'C'
        cell.offset(row=0, column=1).value = 1
    elif cell.offset(row=0,column=-2).value == "Home Landline":
        cell.value = 'H'
        cell.offset(row=0, column=1).value = 1
    else:
        cell.value = 'H'
        cell.offset(row=0, column=1).value = 1





categorize_emails(ws3, 'S', 'T2:T1048576', 'U2:U1048576')


"""Formats the date column to the proper LINKS format"""
for row in ws3.iter_rows(min_row=2, min_col=column_index_from_string('V'), max_col=column_index_from_string('V')):
    for cell in row:
        try:
            cell.number_format = 'General'
            cell.value = int(cell.value.strftime("%Y%m%d"))
        except:
            continue

"""Deletes all rows that don't have a address"""
for cell in ws3['E']:
    if cell.value is None:
        ws3.delete_rows(cell.row,1)


"""Creates a title row in Campaigner Contact Update"""
title_row = ["LOOKUP ID", "FIRST_NAME", "MIDDLE_NAME",	 "LAST_NAME",	"Street1", "Street2", "Street3",	"Street4",
            "CITY", "STATE","Postal_Code", "COUNTRY", "Address Type", "Address is Primary",  "Preferred Home" , "Phone",
            "Phone Type", "Phone is Primary", "Email", "Email Type", "Email is Primary",	"Last_UPDT", "Source"]
i = 0
for row in ws3.iter_rows(min_row=1, max_row=1, max_col=len(title_row)):
    for cell in row:
        cell.value = title_row[i]
        cell.font = Font(bold=True)
        i = i + 1

"""Creates data validation drop down lists for campaigner contact update template"""
dv1 = DataValidation(type="list", formula1='"H,B,A,O,P,S"', allow_blank=True)
create_data_validation(dv1, ws3, 'M')

dv2 = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
create_data_validation(dv2, ws3, 'N')
create_data_validation(dv2, ws3, 'R')
create_data_validation(dv2, ws3, 'U')

dv3 = DataValidation(type="list", formula1='"H,B,C,F,0,S"', allow_blank=True)
create_data_validation(dv3, ws3, 'Q')

dv4 = DataValidation(type="list", formula1='"H,A,B,P"', allow_blank=True)
create_data_validation(dv4, ws3, 'T')

# This drop down is ready to go if Lynda wants it
# dv5 = DataValidation(type="list", formula1='"Online Contact Update"', allow_blank=True)
# create_date_validation(dv5, ws3, 'W')



values_for_initium = ['B','J', 'K', 'L', 'M','N', 'O', 'I']
"""Copies out Canadian address from Campaigner_workbook and puts them into a Initium ready file"""
for cell in ws1['I']:
    if cell.value == 'Canada' and ws1.cell(row=cell.row, column=2).value != "unable to locate":
        all_info = []
        for key in values_for_initium:
            all_info.append(ws1.cell(row=cell.row, column=column_index_from_string(key)).value)
        ws4.append(all_info)
    elif cell.value == 'United States Of America' and ws1.cell(row=cell.row, column=2).value != "unable to locate":
        all_info = []
        for key in values_for_initium:
            all_info.append(ws1.cell(row=cell.row, column=column_index_from_string(key)).value)
        ws6.append(all_info)


convert_from_string_to_int(ws4, 'A')
convert_from_string_to_int(ws6, 'A')

delete_empty_rows(ws4, 'B')
delete_empty_rows(ws6, 'B')


"""Creates a title row for Initium_Ready file"""
initium_title_row = ["LOOKUP ID", "Street1", "Street2", "Street3", "Street4", "CITY",
             "STATE", "Postal_Code", "COUNTRY"]

ws4.insert_rows(1, 1)
ws4.insert_cols(5, 1)
ws6.insert_rows(1, 1)
ws6.insert_cols(5, 1)


i=0
for row in ws4.iter_rows(min_row=1, max_row=1, max_col=9):
    for cell in row:
        cell.value = initium_title_row[i]
        cell.font = Font(bold=True, color='FF0000')
        i = i + 1
i=0
for row in ws6.iter_rows(min_row=1, max_row=1, max_col=9):
    for cell in row:
        cell.value = initium_title_row[i]
        cell.font = Font(bold=True, color='FF0000')
        i = i + 1


"""Creates workbook for the business information based on the campaigner_workbook"""
wb5 = Workbook()
ws5 = wb5.active


business_file_titles_row = ["LOOKUPID", "First Name", "Middle Name", "Last Name", "UBC Degree",
                            "Graduation Year (most recent)", "Job Title", "Company Name", "Address 1", "Address 2",	"Address 3",
                            "Address 4","City","Province","Postal code","Country","Business Phone", "Business Email"]


business_file_needed_column = ['B','D','E','F','G','H','T', 'U', 'X','Y','Z','AA','AB','AD','AN','V','AQ','AS']

for cell in ws1['S']:
    if cell.value is not None:
        one_row = []
        for key in business_file_needed_column:
            one_row.append(ws1.cell(row=cell.row, column=column_index_from_string(key)).value)
        ws5.append(one_row)

"""Converts most LOOKUPID's back to integers to prevent warnings in excel (ex "this number is stored as string"""
for cell in ws5['A']:
    try:
        cell.value = int(cell.value)
    except:
        continue


"""Combines street address 1 and 2 into one with the Business Info Sheet"""
for cell in ws5['J']:
    if cell.offset(row=0, column=-1).value is not None:
        cell_tuple = (str(cell.offset(row=0, column=-1).value), str(cell.value))
        x = "-".join(cell_tuple)
        cell.value = x
        cell.offset(row=0, column=-1).value = ""
    try:
        if cell.offset(row=0, column=2).value.lower() == cell.offset(row=0,column=3).value.lower():
            cell.offset(row=0, column=2).value = ""
    except:
        continue
ws5.delete_cols(column_index_from_string('I'),1)
ws5.insert_cols(column_index_from_string('K'), 1)

i=0
for row in ws5.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.value = business_file_titles_row[i]
        cell.font = Font(bold=True, color='FF0000')
        i = i + 1

wb3.save("Campaigner - Contact_Update_Template.xlsx")
wb4.save("CANADA Initium Ready.xlsx")
wb5.save("Business Addresss from Campaigner.xlsx")
wb6.save("USA Initium Ready.xlsx")














