from openpyxl import load_workbook, Workbook
import glob
from myFunctions import *
from stateAbbreviations import *
import os, sys
import unicodedata

"""Campainger1 will move the necessary cells in the large Campaigner file to something more succient and somewhat
smaller. It also creates a file (cmt.xlsx) which is ready for the Link Constituent Matching Tool"""

os.chdir("C:/Users/sakiikas/Documents/ScriptFiles_TEST/Campaigner")

#os.chdir("W:/Records/LyndaScript-master/Campaigner Files")


Campaigner_download_file = (glob.glob("*Download*")[0])

workbook1 = Campaigner_download_file

wb1 = load_workbook(workbook1)
ws1 = wb1.active

"""Moves values over from K column and places them into Q column"""
for row in ws1.iter_rows(min_col=column_index_from_string('K'),max_col=column_index_from_string('K'), min_row=2):
    for cell in row:
        if cell.value is not None:
            cell.offset(row=0, column=6).value = cell.value
            cell.value = None

"""Moves values over from P column and places them into Q column"""
for row in ws1.iter_rows(min_col=column_index_from_string('P'),max_col=column_index_from_string('P'), min_row=2):
    for cell in row:
        if cell.value is not None:
            cell.offset(row=0, column=1).value = cell.value
            cell.value = None

"""Moves values over from R S T U V W X Y Z columns and places them into Q column"""
for row in ws1.iter_rows(min_col=column_index_from_string('R'),max_col=column_index_from_string('Z'), min_row=2):
    for cell in row:
        if cell.value is not None:
            ws1.cell(row=cell.row, column=column_index_from_string('Q')).value = cell.value
            cell.value = None

"""Moves values over from AB and AC columns and places them into AA column"""
for row in ws1.iter_rows(min_col=column_index_from_string('AB'),max_col=column_index_from_string('AC'), min_row=2):
    for cell in row:
        if cell.value is not None:
            ws1.cell(row=cell.row, column=column_index_from_string('AA')).value = cell.value
            cell.value = None

"""Moves values over from AS,AT,AU,AV,AW,AX,AY,AZ,BA and places them into AR column ( Business Provinces)"""
for row in ws1.iter_rows(min_col=column_index_from_string('AS'),max_col=column_index_from_string('BA'), min_row=2):
    for cell in row:
        if cell.value is not None:
            ws1.cell(row=cell.row, column=column_index_from_string('AR')).value = cell.value
            cell.value = None

"""Moves values over from AQ and places them into AR column (Business Provinces)"""
for row in ws1.iter_rows(min_col=column_index_from_string('AQ'),max_col=column_index_from_string('AQ'), min_row=2):
    for cell in row:
        if cell.value is not None:
            ws1.cell(row=cell.row, column=column_index_from_string('AR')).value = cell.value
            cell.value = None

"""Moves values over from AK and places them into AR column (Business Provinces)"""
for row in ws1.iter_rows(min_col=column_index_from_string('AK'),max_col=column_index_from_string('AK'), min_row=2):
    for cell in row:
        if cell.value is not None:
            ws1.cell(row=cell.row, column=column_index_from_string('AR')).value = cell.value
            cell.value = None

"""Moves values over from BC,BD  and places them into BB column (Business Postal Codes)"""
for row in ws1.iter_rows(min_col=column_index_from_string('BC'),max_col=column_index_from_string('BD'), min_row=2):
    for cell in row:
        if cell.value is not None:
            ws1.cell(row=cell.row, column=column_index_from_string('BB')).value = cell.value
            cell.value = None



wb1.save("C:/Users/sakiikas/Documents/ScriptFiles_TEST/Campaigner/Campaigner_Workbook.xlsx")

# wb1.save("W:/Records/LyndaScript-master/Campaigner Files/Campaigner_Workbook.xlsx")


workbook2 = ('Campaigner_Workbook.xlsx')
wb2 = load_workbook(workbook2)
ws2 = wb2.active


"""Formats phone number by removing extra spaces and unnecessary characters"""
for row in ws2.iter_rows(min_row=2, min_col=column_index_from_string('AD'), max_col=column_index_from_string('AD')):
    for cell in row:
        phone = str(cell.value)
        cell.value = phone.replace('-', '').replace('(', '').replace(')', '').replace(' ', '').\
            replace('None', '').replace('#', '').replace('.', '').replace('+','').replace('=','')
        if len(cell.value) > 10:
            if cell.value.startswith('1') and (cell.offset(row=0, column=-21).value == "Canada" or cell.offset(row=0, column=-21).value == "United States Of America"):
                cell.value = cell.value.replace('1', '')
            else:
                for key in country_codes:
                    if cell.value.startswith(key):
                        if country_codes[key] == cell.offset(row=0, column=-21).value:
                            cell.value = cell.value.replace(key, '')
                            break
                        else:
                            cell.fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
                            break

        try:
            cell.value = int(cell.value)
        except:
            pass
for cell in ws2['AD']:
    if len(str(cell.value)) > 10 and not str(cell.value).startswith('Preferred Phone'):
        cell.fill = PatternFill(fgColor='FDAB9F', fill_type='solid')

"""Formats phone number by removing extra spaces and unnecessary characters for business phone numbers"""
for row in ws2.iter_rows(min_row=2, min_col=column_index_from_string('BE'), max_col=column_index_from_string('BE')):
    for cell in row:
        phone = str(cell.value)
        cell.value = phone.replace('-', '').replace('(', '').replace(')', '').replace(' ', '').replace('None', '')\
            .replace('.', '').replace('+', '').replace('=', '').replace('#', '')
        if cell.value is None or cell.value == '':
            continue
        else:
            try:
                cell.value = int(cell.value)
            except:
                continue

"""Trys to remove accents from First/Middle/Last name, Address1234, City, and Country"""
for row in ws2.iter_rows(min_col=2, max_col=column_index_from_string('BD')):
    for cell in row:
        try:
            cell.value = unidecode.unidecode(cell.value)
        except:
            continue

"""Formats the states of both popular countries, USA, and Canada and formats to their matching state"""
"""There are several cases when the cell will be flagged (highlighted):
-If the state given does not match something in our dictionary (aka mispelled)
-
-If the country os european and a city is written, state can be deleted"""
for cell in ws2['I']:
    if cell.value in popular_countries and cell.offset(row=0, column=8).value is not None:
        try:
            cell.offset(row=0, column=8).value = popular_countries[cell.value][cell.offset(row=0, column=8).value]
        except:
            cell.offset(row=0, column=8).fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
    elif cell.value in usa_canada and cell.offset(row=0, column=8).value is not None:
        try:
            cell.offset(row=0, column=8).value = usa_canada[cell.value][cell.offset(row=0, column=8).value]
        except:
            cell.offset(row=0, column=8).fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
    elif cell.value in european_countries_and_singapore and cell.offset(row=0, column=6).value is not None:
        cell.offset(row=0, column=8).value = ''
    elif cell.offset(row=0, column=3).value is None or cell.offset(row=0, column=8).value == 'Province':
        pass
    # elif cell.offset(row=0, column=8).value is not None and cell.offset(row=0, column=8).value != "Province":
    #     cell.offset(row=0, column=8).fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
    else:
        cell.offset(row=0, column=8).fill = PatternFill(fgColor='FDAB9F', fill_type='solid')
    if cell.offset(row=0, column=6).value is None and (cell.value == 'Singapore' or cell.value == 'Hong Kong'):
        cell.offset(row=0, column=6).fill = PatternFill(fgColor='FDAB9F', fill_type='solid')




"""Combines street address 1 and 2 into one"""
"""Try statements removes extra symbols and words from the address - such as #, commas, Apt"""
for cell in ws2['L']:

    if cell.offset(row=0, column=-2).value is not None:
        cell_tuple = (str(cell.offset(row=0, column=-2).value), str(cell.value))
        x = "-".join(cell_tuple)
        cell.value = x
        cell.offset(row=0, column=-2).value = ""
    if cell.offset(row=0, column=2).value == cell.offset(row=0,column=3).value:
        cell.offset(row=0, column=2).value = ""

    try:
        address1 = str(cell.value).title()
        address2 = str(cell.offset(row=0, column=1).value).title()
        address3 = str(cell.offset(row=0, column=2).value).title()
        cell.value = address1.replace('(', '').replace(')', '').replace('.', ' ').replace('#', ' ').\
            replace(',', ' ').replace('None', '').replace(' - ', '-').replace('- ', '-').replace(' -', '-')\
            .replace('  ', ' ').replace('Th', 'th')
        cell.offset(row=0, column=1).value = address2.replace('(', '').replace(')', '').replace('.', ' ')\
            .replace('#', ' ').replace(',', ' ').replace('None', '').replace('- ', '-').replace(' -', '-')\
            .replace('  ', ' ').replace('Th', 'th')
        cell.offset(row=0, column=2).value = address3.replace('(', '').replace(')', '').replace('.', ' ')\
            .replace('#', ' ').replace(',', ' ').replace('None', '').replace('- ', '-')\
            .replace(' -', '-').replace('  ', ' ').replace('Th', 'th')
        """Takes out the first iteration of Apt or Apartment for Canadian addresss"""

    except Exception as e:
        pass

    for key in unwanted_words:
        if cell.value.startswith(key):
            cell.value = cell.value.replace(key, unwanted_words[key])
            break


    """highlights column b"""
    cell.offset(row=0, column=-10).fill = PatternFill(fgColor='FFFF33', fill_type = 'solid')


format_postal_code(ws2, 'I', 18)

wb2.save("Campaigner_Workbook.xlsx")


wb3 = Workbook()
ws3 = wb3.active


column_list = ['D', 'E', 'F', 'L', 'M', 'N', 'O', 'Q', 'AA', 'I', 'AF', 'AD', 'G']
information_from_excel = list(make_column_list(ws2, column_list))
maximum_rows = len(information_from_excel)
maximum_col = len(information_from_excel[0])



i = 0
for rows in ws3.iter_rows(max_row=maximum_rows, max_col=maximum_col):
    j = 0
    for cell in rows:
        cell.value = information_from_excel[i][j]
        j = j+1
    i = i + 1

ws3.insert_cols(1, 1)
ws3.insert_cols(5, 1)
ws3.insert_cols(9, 1)



cmt_title_row = ["Title", "First name","Middle","Last name", "Suffix", "Address 1", "Address 2",	"Address 3",
                     "Address 4","City","Province","Postal code",	"Country","Email","Phone",	"Degree Info",
                     "Alternate ID type","Alternate ID"]
i = 0
for row in ws3.iter_rows(min_row=1, max_row=1, max_col=len(cmt_title_row)):
    for cell in row:
        cell.value = cmt_title_row[i]
        cell.font = Font(bold=True, color='FF0000')
        i = i + 1

ws2.delete_cols(column_index_from_string('AB'), 2)
ws2.delete_cols(column_index_from_string('R'), 9)
ws2.delete_cols(column_index_from_string('P'), 1)
ws2.delete_cols(column_index_from_string('J'), 2)


wb3.save("cmt.xlsx")
wb2.save("Campaigner_Workbook.xlsx")
